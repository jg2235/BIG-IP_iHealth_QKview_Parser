#!/usr/bin/env python3
"""
ihealth_sizing.py — Extract BIG-IP sizing intelligence from iHealth QKViews.

For an F5 SE building a sizing recommendation. Pulls, per QKView:
  - Identity:     hostname, platform, version, serial number (S/N)
  - Config totals: object counts (virtuals, pools, nodes, monitors, profiles,
                   iRules, SNAT, GTM wide-IPs, etc.) parsed from bigip*.conf
  - Licensing & provisioning: licensed + provisioned modules and levels
  - Performance:  HIGHEST and AVERAGE for each metric over 3h / 1d / 7d / 30d:
                  Blade-0 CPU Usage by Core, Throughput (Bits), SSL Transactions,
                  Active Connections, Memory Used, DNS Requests.

Data sources (current iHealth API, ihealth-api.f5.com):
  - Identity: GET /qkviews/<id>/diagnostics.json?set=hit  (system_information+version)
  - Licensing: /config/bigip.license            via Files API (octet-stream)
  - Provisioning: /config/bigip_base.conf        via Files API
  - Config totals: /config/bigip.conf + bigip_base.conf
  - Performance: /var/tmp/qkview-rrd/*.xml.gz   via Files API,
                 gunzip -> rrdtool restore -> rrdtool graph PRINT VDEF
                 (MAXIMUM from MAX RRA, AVERAGE from AVERAGE RRA, per window)

Auth:
  export IHEALTH_CLIENT_ID=...
  export IHEALTH_CLIENT_SECRET=...

Usage:
  python3 ihealth_sizing.py                 # analyze QKVIEWS, write Excel
  python3 ihealth_sizing.py --discover      # list /var/tmp/qkview-rrd files + DS
  python3 ihealth_sizing.py --upload FILE   # upload a qkview, print its id

Dependencies: requests, openpyxl, rrdtool (apt-get install rrdtool).
"""
from __future__ import annotations

import argparse
import base64
import gzip
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
from collections import OrderedDict
from typing import Any, Dict, List, Optional, Tuple

import requests

# =============================================================================
# INPUT SECTION — edit these
# =============================================================================

QKVIEWS: "OrderedDict[str, str]" = OrderedDict([
    ("2xxxxxx", "device-01.com"),
    ("2xxxxxx", "device-02.com"),
])

OUTPUT_XLSX = "ihealth_sizing_report.xlsx"

# Performance metric -> RRD source mapping (xml.gz file under /var/tmp/qkview-rrd).
# DS names tuned to F5 BIG-IP RRD schema as observed on TMOS 17.x (Z101 platform).
# Run --discover on a different device if DS names look different.
PERF_METRICS: "OrderedDict[str, Dict[str, Any]]" = OrderedDict([
    ("CPU Usage by Core (%)", {
        # Match all blade CPU RRDs: blade0cpu.xml.gz (appliance / slot-0
        # blade) AND blade1cpu.xml.gz / blade2cpu.xml.gz / ... (chassis).
        "files_regex": r"^blade\d+cpu\.xml\.gz$",
        "file_label_regex": r"^blade(\d+)cpu",
        "file_label_format": "Blade {}",
        # blade<N>cpu DSes are per-core jiffy-rate counters
        # (S<slot>C<core>{idle,user,system,iowait,irq,softirq,niced,stolen}).
        # CPU usage % is computed in rrdtool CDEF as
        #   (sum(all modes) - idle) / sum(all modes) * 100
        # which is self-normalizing across kernel HZ values.
        "compute": "cpu_usage",
        "label_regex": r"S\d+C(\d+)idle",
        "label_format": "Core {}",
        "unit": "%",
    }),
    ("Throughput (Bits)", {
        "files": ["throughput.xml.gz"],
        # Match the iHealth GUI "Throughput" graph's *Service* series
        # (service-side throughput), not in+out sum. Priority groups:
        # first group with any DS hit wins. Group 1 catches service DS
        # variants (tput_svc_bytes / svcbytes / service_bytes); group 2
        # falls back to in+out sum only if no service DS exists in this
        # TMOS version's RRD schema. Bytes/s counters -> x8 = bits/s.
        "ds_match_priority": [["svc", "service"], ["tput_bytes"]],
        "aggregate": "sum",
        "scale": 8.0, "unit": "bits/s",
    }),
    ("SSL Transactions", {
        "files": ["connections.xml.gz"],
        # Client-side SSL transaction counters (Compatible mode + Native mode).
        # RRD COUNTER -> rrdtool yields per-second rates automatically.
        "ds_match": ["ssltotcomclient", "ssltotnatclient"],
        "aggregate": "sum",
        "scale": 1.0, "unit": "TPS",
    }),
    ("Active Connections", {
        "files": ["connections.xml.gz"],
        # Substring "curclientconns" matches curclientconns + pvacurclientconns
        # (slow-path + PVA fast-path); sum = total current client connections.
        "ds_match": ["curclientconns"],
        "aggregate": "sum",
        "scale": 1.0, "unit": "conns",
    }),
    ("Memory Used", {
        "files": ["memory.xml.gz"],
        # Use the rollup DSes (R-prefixed) = device-wide totals across all
        # populated blade slots. Works for both appliance and chassis without
        # double-counting per-blade B<n> series.
        "ds_match": ["Rtmmused", "Rotherused"],
        "aggregate": "sum",
        "scale": 1.0, "unit": "bytes",
    }),
    ("Memory Used (%)", {
        "files": ["memory.xml.gz"],
        # Computed via rrdtool CDEF: (Rtmmused + Rotherused) / Rtotal * 100.
        # Matches what the iHealth GUI graph "Memory Used" displays on the
        # Percent Used axis (TMM + Other, excludes swap).
        "compute": "percent",
        "numerator_ds":   ["Rtmmused", "Rotherused"],
        "denominator_ds": ["Rtotal"],
        "unit": "%",
    }),
    ("DNS Requests", {
        # Standard DNS RRDs ship only when GTM/DNS is provisioned. Will N/A on
        # non-DNS devices, which is the correct/honest result.
        "files": ["dns.xml.gz", "gtm.xml.gz", "dnsx.xml.gz"],
        "ds_match": ["request", "query", "dns"],
        "aggregate": "sum",
        "scale": 1.0, "unit": "req/s",
    }),
])

INTERVALS: List[Tuple[str, str]] = [
    ("3h", "3 hour"), ("1d", "1 day"), ("7d", "7 day"), ("30d", "30 day"),
]

CONFIG_COUNTS: "OrderedDict[str, str]" = OrderedDict([
    ("LTM Virtual Servers",  "ltm virtual "),
    ("LTM Pools",            "ltm pool "),
    ("LTM Nodes",            "ltm node "),
    ("LTM Monitors",         "ltm monitor "),
    ("LTM Profiles",         "ltm profile "),
    ("LTM SNAT/SNATPool",    "ltm snat"),
    ("LTM iRules",           "ltm rule "),
    ("LTM Policies",         "ltm policy "),
    ("Persistence Profiles", "ltm persistence "),
    ("GTM Wide IPs",         "gtm wideip "),
    ("GTM Pools",            "gtm pool "),
    ("GTM Servers",          "gtm server "),
    ("APM Access Profiles",  "apm profile access "),
    ("ASM Policies",         "asm policy "),
    ("Net Self IPs",         "net self "),
    ("Net VLANs",            "net vlan "),
    ("Net Routes",           "net route "),
])

# =============================================================================
# iHealth API client
# =============================================================================

API_BASE = os.environ.get("IHEALTH_API_BASE",
                          "https://ihealth-api.f5.com/qkview-analyzer/api")
TOKEN_URL = os.environ.get(
    "IHEALTH_TOKEN_URL",
    "https://identity.account.f5.com/oauth2/ausp95ykc80HOU7SQ357/v1/token",
)
ACCEPT_JSON = "application/vnd.f5.ihealth.api+json"
ACCEPT_OCTET = "application/octet-stream"
USER_AGENT = os.environ.get("IHEALTH_UA", "ihealth-sizing/1.1")


class IHealthError(RuntimeError):
    pass


class IHealthClient:
    def __init__(self) -> None:
        self.s = requests.Session()
        self.s.headers["User-Agent"] = USER_AGENT
        self._token: Optional[str] = os.environ.get("IHEALTH_TOKEN")
        self._acquired = time.time() if self._token else 0.0

    def authenticate(self) -> None:
        cid = os.environ.get("IHEALTH_CLIENT_ID")
        sec = os.environ.get("IHEALTH_CLIENT_SECRET")
        if not (cid and sec):
            raise IHealthError("Set IHEALTH_CLIENT_ID and IHEALTH_CLIENT_SECRET")
        basic = base64.b64encode(f"{cid}:{sec}".encode()).decode()
        r = requests.post(TOKEN_URL, headers={
            "accept": "application/json",
            "authorization": f"Basic {basic}",
            "cache-control": "no-cache",
            "content-type": "application/x-www-form-urlencoded",
        }, data="grant_type=client_credentials&scope=ihealth", timeout=30)
        r.raise_for_status()
        self._token = r.json().get("access_token")
        if not self._token:
            raise IHealthError(f"No access_token: {r.text[:200]}")
        self._acquired = time.time()

    def _auth(self) -> Dict[str, str]:
        if not self._token or (time.time() - self._acquired) > 1500:
            self.authenticate()
        return {"Authorization": f"Bearer {self._token}"}

    def _get(self, path: str, accept: str = ACCEPT_JSON,
             max_tries: int = 12, sleep_s: float = 10.0) -> requests.Response:
        url = f"{API_BASE}{path}"
        backoff = 3.0
        for attempt in range(1, max_tries + 1):
            try:
                r = self.s.get(url, headers={"Accept": accept, **self._auth()},
                               allow_redirects=False, timeout=180)
            except (requests.exceptions.ConnectionError,
                    requests.exceptions.Timeout) as e:
                if attempt < max_tries:
                    print(f"  net err ({attempt}/{max_tries}): {e}; "
                          f"retrying in {backoff:.0f}s", file=sys.stderr)
                    time.sleep(backoff); backoff = min(backoff * 2, 60); continue
                raise
            if r.status_code == 202 and attempt < max_tries:
                print(f"  202 processing ({attempt}/{max_tries}); "
                      f"waiting {sleep_s}s", file=sys.stderr)
                time.sleep(sleep_s); continue
            if r.status_code in (500, 502, 503, 504) and attempt < max_tries:
                print(f"  HTTP {r.status_code} transient ({attempt}/{max_tries}); "
                      f"retrying in {backoff:.0f}s", file=sys.stderr)
                time.sleep(backoff); backoff = min(backoff * 2, 60); continue
            return r
        return r

    def diagnostics(self, qid: str) -> Dict[str, Any]:
        r = self._get(f"/qkviews/{qid}/diagnostics.json?set=hit")
        if r.status_code != 200:
            raise IHealthError(f"diagnostics {qid}: HTTP {r.status_code}: {r.text[:200]}")
        return r.json()

    def list_files(self, qid: str) -> List[Tuple[str, str]]:
        r = self._get(f"/qkviews/{qid}/files")
        if r.status_code != 200:
            raise IHealthError(f"files {qid}: HTTP {r.status_code}: {r.text[:200]}")
        body = r.json()
        out: List[Tuple[str, str]] = []
        if isinstance(body, list):
            for f in body:
                fid = f.get("id")
                path = f.get("value") or f.get("path") or f.get("name")
                if fid and path:
                    out.append((str(fid), str(path)))
        elif isinstance(body, dict):
            for _k, v in body.items():
                if isinstance(v, list):
                    for it in v:
                        if isinstance(it, dict):
                            fid = it.get("id") or it.get("@id")
                            path = (it.get("value") or it.get("path")
                                    or it.get("name") or it.get("$")
                                    or it.get("#text"))
                            if fid and path:
                                out.append((str(fid), str(path)))
        return out

    def download_file(self, qid: str, file_hash: str, out_path: str) -> bool:
        r = self._get(f"/qkviews/{qid}/files/{file_hash}", accept=ACCEPT_OCTET)
        if r.status_code != 200:
            return False
        with open(out_path, "wb") as fh:
            fh.write(r.content)
        return True

    def upload(self, path: str, description: Optional[str] = None,
               case: Optional[str] = None) -> Optional[str]:
        data: Dict[str, str] = {"visible_in_gui": "true"}
        if description: data["description"] = description
        if case: data["f5_support_case"] = case
        with open(path, "rb") as fh:
            r = self.s.post(f"{API_BASE}/qkviews",
                            headers={"Accept": ACCEPT_JSON, **self._auth()},
                            files={"qkview": fh}, data=data,
                            allow_redirects=False, timeout=600)
        loc = r.headers.get("Location", "")
        print(f"HTTP {r.status_code} | body: {r.text[:200]}", file=sys.stderr)
        return loc.rstrip("/").split("/")[-1] if loc else None


# =============================================================================
# Identity (diagnostics)
# =============================================================================

def _walk_pairs(obj: Any):
    if isinstance(obj, dict):
        for k, v in obj.items():
            yield k, v
            yield from _walk_pairs(v)
    elif isinstance(obj, list):
        for it in obj:
            yield from _walk_pairs(it)


def _find_first(obj: Any, key: str) -> Optional[Any]:
    for k, v in _walk_pairs(obj):
        if k == key:
            return v
    return None


def _scalar(v: Any) -> Optional[str]:
    if v is None: return None
    if isinstance(v, (str, int, float)):
        s = str(v).strip()
        return s or None
    if isinstance(v, dict):
        for k in ("$", "#text", "value"):
            if k in v:
                return _scalar(v[k])
    return None


def extract_identity(client: IHealthClient, qid: str) -> Dict[str, str]:
    info = {"Hostname": "N/A", "Platform": "N/A", "Version": "N/A",
            "Product": "N/A", "Build": "N/A", "Edition": "N/A",
            "Chassis S/N": "N/A"}
    try:
        d = client.diagnostics(qid)
    except IHealthError as e:
        print(f"  identity warning: {e}", file=sys.stderr)
        return info
    info["Hostname"] = _scalar(_find_first(d, "hostname")) or info["Hostname"]
    info["Platform"] = _scalar(_find_first(d, "platform")) or info["Platform"]
    info["Chassis S/N"] = (_scalar(_find_first(d, "bigip_chassis_serial_num"))
                           or info["Chassis S/N"])
    vraw = _find_first(d, "version")
    if isinstance(vraw, dict):
        info["Version"] = _scalar(vraw.get("version")) or info["Version"]
        info["Product"] = _scalar(vraw.get("product")) or info["Product"]
        info["Build"]   = _scalar(vraw.get("built"))   or info["Build"]
        info["Edition"] = _scalar(vraw.get("edition")) or info["Edition"]
    else:
        info["Version"] = _scalar(vraw) or info["Version"]
        info["Product"] = _scalar(_find_first(d, "product")) or info["Product"]
        info["Build"]   = _scalar(_find_first(d, "built"))   or info["Build"]
        info["Edition"] = _scalar(_find_first(d, "edition")) or info["Edition"]
    return info


# =============================================================================
# License + Provisioning
# =============================================================================

def _find_file_id(files: List[Tuple[str, str]], path: str) -> Optional[str]:
    for fid, p in files:
        if p == path:
            return fid
    return None


def parse_license_text(text: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {"active_modules": [], "platform_id": "N/A",
                           "registration_key": "N/A", "service_check_date": "N/A",
                           "license_type": "N/A"}
    for line in text.splitlines():
        line = line.rstrip()
        if line.startswith("#") or ":" not in line:
            continue
        key, _, val = line.partition(":")
        key = key.strip().lower(); val = val.strip()
        if key == "active module":
            first = val.split("|", 1)[0]
            out["active_modules"].append(first.strip())
        elif key == "registration key":
            out["registration_key"] = val
        elif key == "service check date":
            out["service_check_date"] = val
        elif key == "platform id":
            out["platform_id"] = val
        elif key == "license type":
            out["license_type"] = val
    return out


_PROV_BLOCK = re.compile(r"^sys\s+provision\s+(\S+)\s*\{(.*?)^\}",
                         re.M | re.S)
_LEVEL_LINE = re.compile(r"^\s*level\s+(\S+)", re.M)


def parse_provisioning_text(text: str) -> List[Tuple[str, str]]:
    out: List[Tuple[str, str]] = []
    for m in _PROV_BLOCK.finditer(text):
        module = m.group(1); body = m.group(2)
        lv = _LEVEL_LINE.search(body)
        level = lv.group(1) if lv else "none"
        if level.lower() != "none":
            out.append((module, level))
    return out


def extract_license_provisioning(client: IHealthClient, qid: str,
                                 files: List[Tuple[str, str]]) -> Dict[str, Any]:
    res: Dict[str, Any] = {"provisioned": [], "active_modules": [], "headline": {}}
    with tempfile.TemporaryDirectory() as td:
        lic_id = _find_file_id(files, "/config/bigip.license")
        if lic_id:
            p = os.path.join(td, "bigip.license")
            if client.download_file(qid, lic_id, p):
                with open(p, encoding="utf-8", errors="replace") as fh:
                    parsed = parse_license_text(fh.read())
                res["active_modules"] = parsed["active_modules"]
                res["headline"] = {
                    "Registration Key": parsed["registration_key"],
                    "Service Check Date": parsed["service_check_date"],
                    "Platform ID": parsed["platform_id"],
                    "License Type": parsed["license_type"],
                }
        base_id = _find_file_id(files, "/config/bigip_base.conf")
        if base_id:
            p = os.path.join(td, "bigip_base.conf")
            if client.download_file(qid, base_id, p):
                with open(p, encoding="utf-8", errors="replace") as fh:
                    res["provisioned"] = parse_provisioning_text(fh.read())
    return res


# =============================================================================
# Config totals
# =============================================================================

def extract_config_totals(client: IHealthClient, qid: str,
                          files: List[Tuple[str, str]]) -> Dict[str, int]:
    counts = OrderedDict((k, 0) for k in CONFIG_COUNTS)
    targets = [("/config/bigip.conf", "bigip.conf"),
               ("/config/bigip_base.conf", "bigip_base.conf")]
    with tempfile.TemporaryDirectory() as td:
        for src, name in targets:
            fid = _find_file_id(files, src)
            if not fid: continue
            p = os.path.join(td, name)
            if not client.download_file(qid, fid, p): continue
            with open(p, encoding="utf-8", errors="replace") as fh:
                for line in fh:
                    if not line or line[0].isspace():
                        continue
                    for label, prefix in CONFIG_COUNTS.items():
                        if line.startswith(prefix):
                            counts[label] += 1; break
    return counts


# =============================================================================
# Performance (xml.gz -> rrdtool restore -> windowed MAX/AVG)
# =============================================================================

def rrd_available() -> bool:
    return shutil.which("rrdtool") is not None


def rrd_restore_xml_gz(xml_gz_path: str, dest_rrd: str) -> bool:
    try:
        xml_path = dest_rrd + ".xml"
        with gzip.open(xml_gz_path, "rb") as gi, open(xml_path, "wb") as go:
            shutil.copyfileobj(gi, go)
        subprocess.run(["rrdtool", "restore", "-f", xml_path, dest_rrd],
                       capture_output=True, check=True)
        os.unlink(xml_path)
        return True
    except (OSError, subprocess.CalledProcessError) as e:
        print(f"  rrd restore failed for {os.path.basename(xml_gz_path)}: {e}",
              file=sys.stderr)
        return False


def rrd_ds_names(path: str) -> List[str]:
    try:
        out = subprocess.run(["rrdtool", "info", path], capture_output=True,
                             text=True, check=True).stdout
    except (OSError, subprocess.CalledProcessError):
        return []
    return sorted(set(re.findall(r"^ds\[([^\]]+)\]", out, re.M)))


def rrd_info_meta(path: str) -> Tuple[Optional[int], set]:
    """Return (last_update_epoch, {RRA CFs present}) from `rrdtool info`.

    QKView RRDs end at capture time, NOT at script runtime. Every window
    query must be anchored to last_update or short windows (3h/1d/7d) fall
    entirely after the data and return NaN.
    """
    try:
        out = subprocess.run(["rrdtool", "info", path], capture_output=True,
                             text=True, check=True).stdout
    except (OSError, subprocess.CalledProcessError):
        return None, set()
    lu = None
    m = re.search(r"^last_update\s*=\s*(\d+)", out, re.M)
    if m:
        lu = int(m.group(1))
    cfs = set(re.findall(r'^rra\[\d+\]\.cf\s*=\s*"([^"]+)"', out, re.M))
    return lu, cfs


def _window_range(window: str, end_epoch: Optional[int]) -> Tuple[str, str]:
    """Build (--start, --end) args anchored at the RRD's last_update."""
    end = str(end_epoch) if end_epoch else "now"
    return f"end-{window}", end


def rrd_window_value(path: str, ds: str, mode: str, window: str,
                     end_epoch: Optional[int] = None,
                     rra_cfs: Optional[set] = None) -> Optional[float]:
    """mode:
       'MAX'        = MAX RRA + VDEF MAXIMUM   -> peak value in window
       'AVERAGE'    = AVERAGE RRA + VDEF AVERAGE -> mean value in window
       'MIN_OF_AVG' = AVERAGE RRA + VDEF MINIMUM -> lowest period-average
                      (used for transformed metrics like CPU usage from idle)
    """
    spec = {"MAX": ("MAX", "MAXIMUM"),
            "AVERAGE": ("AVERAGE", "AVERAGE"),
            "MIN_OF_AVG": ("AVERAGE", "MINIMUM")}
    rra, vdef = spec[mode]
    # F5 qkview RRDs often ship AVERAGE-only RRAs. If a MAX RRA is absent,
    # fall back to peak-of-AVERAGE (VDEF MAXIMUM over the AVERAGE RRA) —
    # the best available "highest" and what the iHealth GUI plots anyway.
    if rra == "MAX" and rra_cfs is not None and "MAX" not in rra_cfs:
        rra, vdef = "AVERAGE", "MAXIMUM"
    start, end = _window_range(window, end_epoch)
    cmd = ["rrdtool", "graph", os.devnull,
           "--start", start, "--end", end,
           f"DEF:x={path}:{ds}:{rra}",
           f"VDEF:v=x,{vdef}", "PRINT:v:%.6lf"]
    try:
        out = subprocess.run(cmd, capture_output=True, text=True,
                             check=True).stdout
    except (OSError, subprocess.CalledProcessError):
        return None
    val = (out.strip().splitlines() or [""])[-1].strip()
    try:
        f = float(val)
        if f != f or f in (float("inf"), float("-inf")):
            return None
        return f
    except ValueError:
        return None


def rrd_cpu_usage_pct(path: str, idle_ds: str, all_modes_ds: List[str],
                      window: str, mode: str,
                      end_epoch: Optional[int] = None) -> Optional[float]:
    """Compute per-core CPU usage % via rrdtool CDEF:
         usage_pct = (sum(all modes) - idle) / sum(all modes) * 100

    F5 stores CPU DSes as per-second jiffy rates (idle/user/system/iowait/
    irq/softirq/niced/stolen), NOT as direct percentages. Their absolute
    values depend on kernel HZ (100/250/1000), so percentage MUST be derived
    by normalizing against the sum of all modes for that core.

    mode='high' -> VDEF MAXIMUM of the per-step usage_pct series (peak load);
    mode='avg'  -> VDEF AVERAGE                                  (mean load).
    """
    start, end = _window_range(window, end_epoch)
    args = ["rrdtool", "graph", os.devnull,
            "--start", start, "--end", end,
            f"DEF:idle={path}:{idle_ds}:AVERAGE"]
    mode_names: List[str] = []
    for i, ds in enumerate(all_modes_ds):
        n = f"m{i}"; mode_names.append(n)
        args.append(f"DEF:{n}={path}:{ds}:AVERAGE")
    # total = m0 + m1 + ... (RPN)
    total_rpn = mode_names[0]
    for n in mode_names[1:]:
        total_rpn += f",{n},+"
    args.append(f"CDEF:total={total_rpn}")
    # usage_pct = (total - idle) / total * 100
    args.append("CDEF:usage_pct=total,idle,-,total,/,100,*")
    vdef = "MAXIMUM" if mode == "high" else "AVERAGE"
    args.append(f"VDEF:v=usage_pct,{vdef}")
    args.append("PRINT:v:%.6lf")
    try:
        out = subprocess.run(args, capture_output=True, text=True,
                             check=True).stdout
    except (OSError, subprocess.CalledProcessError):
        return None
    val = (out.strip().splitlines() or [""])[-1].strip()
    try:
        f = float(val)
        if f != f or f in (float("inf"), float("-inf")):
            return None
        return f
    except ValueError:
        return None


def rrd_percent_value(path: str, num_dses: List[str], den_dses: List[str],
                      window: str, mode: str,
                      end_epoch: Optional[int] = None) -> Optional[float]:
    """Compute (sum(num_dses) / sum(den_dses)) * 100 over a window using
    rrdtool CDEF. mode='high' -> VDEF MAXIMUM of the per-step percent series;
    mode='avg' -> VDEF AVERAGE. Both pull from each DS's AVERAGE RRA so the
    sum is taken at matching timestamps (no co-occurrence assumption errors).
    """
    start, end = _window_range(window, end_epoch)
    args = ["rrdtool", "graph", os.devnull,
            "--start", start, "--end", end]
    num_names: List[str] = []
    for i, ds in enumerate(num_dses):
        n = f"n{i}"; num_names.append(n)
        args.append(f"DEF:{n}={path}:{ds}:AVERAGE")
    den_names: List[str] = []
    for i, ds in enumerate(den_dses):
        n = f"d{i}"; den_names.append(n)
        args.append(f"DEF:{n}={path}:{ds}:AVERAGE")
    # Build RPN: a + b + c = "a,b,+,c,+"
    def rpn_sum(names: List[str]) -> str:
        expr = names[0]
        for n in names[1:]:
            expr += f",{n},+"
        return expr
    num_rpn = rpn_sum(num_names)
    den_rpn = rpn_sum(den_names)
    args.append(f"CDEF:pct={num_rpn},{den_rpn},/,100,*")
    vdef = "MAXIMUM" if mode == "high" else "AVERAGE"
    args.append(f"VDEF:v=pct,{vdef}")
    args.append("PRINT:v:%.6lf")
    try:
        out = subprocess.run(args, capture_output=True, text=True,
                             check=True).stdout
    except (OSError, subprocess.CalledProcessError):
        return None
    val = (out.strip().splitlines() or [""])[-1].strip()
    try:
        f = float(val)
        if f != f or f in (float("inf"), float("-inf")):
            return None
        return f
    except ValueError:
        return None


def _ds_label(ds: str, spec: Dict[str, Any]) -> str:
    """Derive a clean per-series label from a DS name using optional regex+format."""
    rgx = spec.get("label_regex")
    fmt = spec.get("label_format", "{}")
    if rgx:
        m = re.search(rgx, ds)
        if m:
            return fmt.format(*m.groups())
    return ds


def _file_label(path: str, spec: Dict[str, Any]) -> str:
    """Derive a per-file label from a path using file_label_regex+format
    (e.g. 'blade2cpu.xml.gz' -> 'Blade 2'). Returns '' if no regex configured."""
    rgx = spec.get("file_label_regex")
    fmt = spec.get("file_label_format", "{}")
    if rgx:
        m = re.search(rgx, os.path.basename(path))
        if m:
            return fmt.format(*m.groups())
    return ""


def _apply_transform(v: Optional[float], transform: Optional[str]) -> Optional[float]:
    if v is None or not transform:
        return v
    if transform == "100_minus":
        return 100.0 - v
    return v


def select_ds(all_ds: List[str], ds_match: List[str]) -> List[str]:
    if "*" in ds_match:
        return all_ds
    chosen: List[str] = []
    for m in ds_match:
        ml = m.lower()
        for d in all_ds:
            if ml in d.lower() and d not in chosen:
                chosen.append(d)
    return chosen


def select_ds_spec(all_ds: List[str], spec: Dict[str, Any]) -> List[str]:
    """Resolve DS selection: 'ds_match_priority' = ordered fallback groups
    (first group with any match wins); else plain 'ds_match'."""
    groups = spec.get("ds_match_priority")
    if groups:
        for grp in groups:
            hit = select_ds(all_ds, grp)
            if hit:
                return hit
        return []
    return select_ds(all_ds, spec.get("ds_match", []))


def _aggregate(values: List[Optional[float]], mode: str) -> Optional[float]:
    real = [v for v in values if v is not None]
    if not real: return None
    if mode == "sum": return sum(real)
    if mode == "max": return max(real)
    if mode == "first": return real[0]
    return sum(real)


def extract_performance(client: IHealthClient, qid: str,
                        files: List[Tuple[str, str]]
                        ) -> "OrderedDict[str, Any]":
    result: "OrderedDict[str, Any]" = OrderedDict()
    if not rrd_available():
        for metric in PERF_METRICS:
            result[metric] = {"_error": "rrdtool not installed (apt-get install rrdtool)"}
        return result

    # All RRD candidates under /var/tmp/qkview-rrd, indexed by basename.
    rrd_by_name: Dict[str, Tuple[str, str]] = {}
    for fid, p in files:
        if "/var/tmp/qkview-rrd/" in p:
            rrd_by_name[os.path.basename(p)] = (fid, p)

    with tempfile.TemporaryDirectory() as td:
        for metric, spec in PERF_METRICS.items():
            # Resolve which source files to use.
            #   files_regex: every matching file (multi-blade chassis case)
            #   files:       exact filenames, first present wins (single file)
            matched_files: List[Tuple[str, str]] = []
            if spec.get("files_regex"):
                pat = re.compile(spec["files_regex"])
                matched_files = sorted(
                    [(fid, p) for fid, p in rrd_by_name.values()
                     if pat.match(os.path.basename(p))],
                    key=lambda x: x[1])
                tried_label = spec["files_regex"]
            else:
                for fname in spec.get("files", []):
                    if fname in rrd_by_name:
                        matched_files = [rrd_by_name[fname]]
                        break
                tried_label = ", ".join(spec.get("files", []))

            if not matched_files:
                result[metric] = {"_error":
                    f"no qkview-rrd source found (tried: {tried_label})"}
                continue

            series: "OrderedDict[str, Any]" = OrderedDict()
            per_core = spec.get("per_core", False)
            agg_mode = spec.get("aggregate", "sum")
            scale = spec.get("scale", 1.0)
            transform = spec.get("transform")
            # For transformed metrics (CPU = 100 - idle):
            #   highest usage = 100 - lowest period-average idle
            #     -> use MIN_OF_AVG, then transform
            #   average usage = 100 - period-average idle
            #     -> use AVERAGE, then transform
            high_mode = "MIN_OF_AVG" if transform == "100_minus" else "MAX"

            sources: List[str] = []
            errors: List[str] = []

            for fid, src_path in matched_files:
                xml_gz = os.path.join(td, os.path.basename(src_path))
                if not client.download_file(qid, fid, xml_gz):
                    errors.append(f"download failed: {src_path}")
                    continue
                rrd_path = xml_gz + ".rrd"
                if not rrd_restore_xml_gz(xml_gz, rrd_path):
                    errors.append(f"restore failed: {src_path}")
                    continue
                all_ds = rrd_ds_names(rrd_path)
                if not all_ds:
                    errors.append(f"no DS in {os.path.basename(src_path)}")
                    continue

                # Anchor all window queries at the RRD's last data point
                # (qkview capture time), not script runtime; note which
                # RRA CFs exist so 'Highest' can fall back when MAX absent.
                last_up, rra_cfs = rrd_info_meta(rrd_path)

                file_label = _file_label(src_path, spec)

                # ---- 'compute: cpu_usage' — per-core CPU% from jiffy counters ----
                if spec.get("compute") == "cpu_usage":
                    # Each idle DS identifies a core. Sibling mode DSes for
                    # that core share the same prefix (e.g. 'S1C0idle' -> 'S1C0').
                    idle_dses = [d for d in all_ds if d.lower().endswith("idle")]
                    if not idle_dses:
                        errors.append(
                            f"no idle DS in {os.path.basename(src_path)}")
                        continue
                    for idle_ds in idle_dses:
                        prefix = idle_ds[:-len("idle")]  # 'S1C0'
                        core_modes = [d for d in all_ds if d.startswith(prefix)]
                        if idle_ds not in core_modes or len(core_modes) < 2:
                            continue
                        core_label = _ds_label(idle_ds, spec)
                        label = (f"{file_label} / {core_label}"
                                 if file_label else core_label)
                        row: "OrderedDict[str, Any]" = OrderedDict()
                        for win, _ in INTERVALS:
                            hi = rrd_cpu_usage_pct(rrd_path, idle_ds,
                                                   core_modes, win, "high",
                                                   end_epoch=last_up)
                            av = rrd_cpu_usage_pct(rrd_path, idle_ds,
                                                   core_modes, win, "avg",
                                                   end_epoch=last_up)
                            row[win] = {"high": hi, "avg": av}
                        series[label] = row
                    sources.append(
                        f"{src_path} (cpu%: 100*(sum(modes)-idle)/sum(modes))")
                    continue

                # ---- 'compute: percent' shortcut: numerator / denominator ----
                if spec.get("compute") == "percent":
                    num_dses = [d for d in spec["numerator_ds"] if d in all_ds]
                    den_dses = [d for d in spec["denominator_ds"] if d in all_ds]
                    missing_n = set(spec["numerator_ds"]) - set(num_dses)
                    missing_d = set(spec["denominator_ds"]) - set(den_dses)
                    if missing_n or missing_d:
                        errors.append(
                            f"missing DS for percent in {os.path.basename(src_path)}: "
                            f"num={sorted(missing_n)} den={sorted(missing_d)}")
                        continue
                    label = (file_label if (file_label and len(matched_files) > 1)
                             else metric)
                    row = OrderedDict()
                    for win, _ in INTERVALS:
                        hi = rrd_percent_value(rrd_path, num_dses, den_dses,
                                               win, "high", end_epoch=last_up)
                        av = rrd_percent_value(rrd_path, num_dses, den_dses,
                                               win, "avg", end_epoch=last_up)
                        row[win] = {"high": hi, "avg": av}
                    series[label] = row
                    sources.append(
                        f"{src_path} (%, {'+'.join(num_dses)} / {'+'.join(den_dses)})")
                    continue

                ds_matched = select_ds_spec(all_ds, spec)
                if not ds_matched:
                    avail = ", ".join(all_ds[:10]) + ("…" if len(all_ds) > 10 else "")
                    wanted = spec.get("ds_match_priority") or spec.get("ds_match")
                    errors.append(
                        f"no DS matched {wanted} in "
                        f"{os.path.basename(src_path)} (available: {avail})")
                    continue

                if per_core:
                    for ds in ds_matched:
                        ds_label = _ds_label(ds, spec)
                        label = (f"{file_label} / {ds_label}"
                                 if file_label else ds_label)
                        row: "OrderedDict[str, Any]" = OrderedDict()
                        for win, _ in INTERVALS:
                            hi_raw = rrd_window_value(rrd_path, ds, high_mode,
                                                      win, last_up, rra_cfs)
                            av_raw = rrd_window_value(rrd_path, ds, "AVERAGE",
                                                      win, last_up, rra_cfs)
                            hi = _apply_transform(hi_raw, transform)
                            av = _apply_transform(av_raw, transform)
                            row[win] = {"high": None if hi is None else hi * scale,
                                        "avg":  None if av is None else av * scale}
                        series[label] = row
                else:
                    # Aggregate matched DS within this file. If multiple files
                    # matched (e.g. multi-blade), label each row by file.
                    label = (file_label if (file_label and len(matched_files) > 1)
                             else metric)
                    row = OrderedDict()
                    for win, _ in INTERVALS:
                        his_raw = [rrd_window_value(rrd_path, ds, high_mode,
                                                    win, last_up, rra_cfs)
                                   for ds in ds_matched]
                        avs_raw = [rrd_window_value(rrd_path, ds, "AVERAGE",
                                                    win, last_up, rra_cfs)
                                   for ds in ds_matched]
                        his = [_apply_transform(v, transform) for v in his_raw]
                        avs = [_apply_transform(v, transform) for v in avs_raw]
                        hi = _aggregate(his, agg_mode)
                        av = _aggregate(avs, agg_mode)
                        row[win] = {"high": None if hi is None else hi * scale,
                                    "avg":  None if av is None else av * scale}
                    series[label] = row

                sources.append(f"{src_path} (DS: {', '.join(ds_matched)})")

            if not any(k for k in series.keys()):
                result[metric] = {"_error": "; ".join(errors) or "no data"}
                continue

            series["_unit"] = spec.get("unit", "")
            series["_source"] = " | ".join(sources)
            result[metric] = series

    return result


# =============================================================================
# Excel
# =============================================================================

def build_workbook(rows: List[Dict[str, Any]], out_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    FONT = "Arial"
    HEAD = PatternFill("solid", fgColor="1F4E78")
    SUB = PatternFill("solid", fgColor="D9E1F2")
    NA = PatternFill("solid", fgColor="F2F2F2")
    W = Font(name=FONT, bold=True, color="FFFFFF")
    B = Font(name=FONT, bold=True)
    R = Font(name=FONT)
    THIN = Side(style="thin", color="BFBFBF")
    BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CTR = Alignment(horizontal="center", vertical="center")
    LFT = Alignment(horizontal="left", vertical="center")

    def hcell(cell):
        cell.font = W; cell.fill = HEAD; cell.alignment = CTR; cell.border = BORD

    wb = Workbook()

    # ---- Inventory ----
    ws = wb.active; ws.title = "Inventory"
    hdr = ["QKView", "Hostname", "Platform", "Version", "Product",
           "Build", "Edition", "Chassis S/N", "Platform ID (license)"]
    for c, h in enumerate(hdr, 1):
        hcell(ws.cell(row=1, column=c, value=h))
    for i, rr in enumerate(rows, 1):
        idn = rr["identity"]; lic = rr["license_prov"].get("headline", {})
        vals = [rr["label"], idn["Hostname"], idn["Platform"], idn["Version"],
                idn["Product"], idn["Build"], idn["Edition"], idn["Chassis S/N"],
                lic.get("Platform ID", "N/A")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=1 + i, column=c, value=v)
            cell.font = B if c == 1 else R; cell.border = BORD
            cell.alignment = LFT if c <= 2 else CTR
    for c in range(1, len(hdr) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 17
    ws.column_dimensions["B"].width = 30
    ws.freeze_panes = ws.cell(row=2, column=1)

    # ---- Provisioning & License ----
    ws2 = wb.create_sheet("Provisioning & License")
    row = 1
    for rr in rows:
        cell = ws2.cell(row=row, column=1,
            value=f"{rr['label']}  ({rr['identity']['Hostname']})")
        cell.font = B; cell.fill = SUB
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1
        hcell(ws2.cell(row=row, column=1, value="Provisioned Module"))
        hcell(ws2.cell(row=row, column=2, value="Level"))
        row += 1
        prov = rr["license_prov"].get("provisioned") or [("(none)", "")]
        for name, lvl in prov:
            ws2.cell(row=row, column=1, value=name).font = R
            c2 = ws2.cell(row=row, column=2, value=lvl); c2.font = R; c2.alignment = CTR
            row += 1
        row += 1
        hcell(ws2.cell(row=row, column=1, value="Active Licensed Module"))
        hcell(ws2.cell(row=row, column=2, value=""))
        row += 1
        for mod in rr["license_prov"].get("active_modules") or ["(none)"]:
            ws2.cell(row=row, column=1, value=mod).font = R
            row += 1
        row += 1
        for k, v in (rr["license_prov"].get("headline") or {}).items():
            ws2.cell(row=row, column=1, value=k).font = B
            ws2.cell(row=row, column=2, value=v).font = R
            row += 1
        row += 2
    ws2.column_dimensions["A"].width = 70
    ws2.column_dimensions["B"].width = 22

    # ---- Config Totals ----
    ws3 = wb.create_sheet("Config Totals")
    labels = list(CONFIG_COUNTS.keys())
    hcell(ws3.cell(row=1, column=1, value="Config Object"))
    for j, rr in enumerate(rows, 2):
        hcell(ws3.cell(row=1, column=j, value=rr["label"]))
    for i, lab in enumerate(labels, 2):
        c = ws3.cell(row=i, column=1, value=lab); c.font = B; c.border = BORD
        for j, rr in enumerate(rows, 2):
            cell = ws3.cell(row=i, column=j, value=rr["config"].get(lab, 0))
            cell.alignment = CTR; cell.border = BORD; cell.font = R
    trow = len(labels) + 2
    c = ws3.cell(row=trow, column=1, value="TOTAL OBJECTS"); c.font = B; c.fill = SUB
    for j in range(2, len(rows) + 2):
        col = get_column_letter(j)
        cell = ws3.cell(row=trow, column=j, value=f"=SUM({col}2:{col}{trow-1})")
        cell.font = B; cell.fill = SUB; cell.alignment = CTR
    ws3.column_dimensions["A"].width = 26
    for j in range(2, len(rows) + 2):
        ws3.column_dimensions[get_column_letter(j)].width = 22

    # ---- Performance ----
    ws4 = wb.create_sheet("Performance")
    r = 1
    iv_labels = [l for _, l in INTERVALS]
    for rr in rows:
        c = ws4.cell(row=r, column=1,
            value=f"{rr['label']}  ({rr['identity']['Hostname']})")
        c.font = B; c.fill = SUB
        ws4.merge_cells(start_row=r, start_column=1, end_row=r,
                        end_column=2 + 2 * len(iv_labels))
        r += 1
        hcell(ws4.cell(row=r, column=1, value="Metric / Series"))
        hcell(ws4.cell(row=r, column=2, value="Unit"))
        col = 3
        for lbl in iv_labels:
            ws4.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
            hcell(ws4.cell(row=r, column=col, value=lbl))
            ws4.cell(row=r, column=col + 1)
            col += 2
        r += 1
        ws4.cell(row=r, column=1).border = BORD
        ws4.cell(row=r, column=2).border = BORD
        col = 3
        for _ in iv_labels:
            for sub in ("Highest", "Average"):
                c = ws4.cell(row=r, column=col, value=sub)
                c.font = B; c.fill = SUB; c.alignment = CTR; c.border = BORD
                col += 1
        r += 1
        for metric, series in rr["performance"].items():
            if isinstance(series, dict) and "_error" in series:
                ws4.cell(row=r, column=1, value=metric).font = B
                c = ws4.cell(row=r, column=3, value=f"N/A — {series['_error']}")
                c.font = Font(name=FONT, italic=True, color="C00000")
                ws4.merge_cells(start_row=r, start_column=3, end_row=r,
                                end_column=2 + 2 * len(iv_labels))
                for cc in range(1, 3 + 2 * len(iv_labels)):
                    ws4.cell(row=r, column=cc).border = BORD
                r += 1; continue
            unit = series.get("_unit", "")
            mc = ws4.cell(row=r, column=1, value=metric); mc.font = B; mc.fill = SUB
            ws4.merge_cells(start_row=r, start_column=1, end_row=r,
                            end_column=2 + 2 * len(iv_labels))
            r += 1
            for label, row_data in series.items():
                if label in ("_unit", "_source"):
                    continue
                c = ws4.cell(row=r, column=1, value=label); c.font = R
                c.alignment = LFT; c.border = BORD
                c = ws4.cell(row=r, column=2, value=unit); c.font = R
                c.alignment = CTR; c.border = BORD
                col = 3
                for win, _ in INTERVALS:
                    cd = row_data.get(win, {})
                    for key in ("high", "avg"):
                        v = cd.get(key)
                        cell = ws4.cell(row=r, column=col,
                            value=(round(v, 2) if isinstance(v, float) else "N/A"))
                        cell.alignment = CTR; cell.border = BORD; cell.font = R
                        cell.number_format = "#,##0.00"
                        if v is None:
                            cell.fill = NA
                        col += 1
                r += 1
            if series.get("_source"):
                c = ws4.cell(row=r, column=1,
                    value=f"Source: iHealth RRD {series['_source']}")
                c.font = Font(name=FONT, italic=True, size=8, color="808080")
                ws4.merge_cells(start_row=r, start_column=1, end_row=r,
                                end_column=2 + 2 * len(iv_labels))
                r += 1
        r += 2
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 10
    for c in range(3, 3 + 2 * len(iv_labels)):
        ws4.column_dimensions[get_column_letter(c)].width = 14

    # ---- Notes ----
    wsn = wb.create_sheet("Notes")
    notes = [
        "F5 iHealth QKView sizing extract.",
        "",
        "Data sources (current iHealth API at ihealth-api.f5.com):",
        " - Identity: /qkviews/<id>/diagnostics  (system_information + version)",
        " - License: /config/bigip.license       via Files API",
        " - Provisioning: /config/bigip_base.conf via Files API",
        " - Config totals: /config/bigip.conf + bigip_base.conf",
        " - Performance: /var/tmp/qkview-rrd/*.xml.gz (gunzip + rrdtool restore)",
        "",
        "Performance method:",
        " - All windows are anchored at the RRD's last_update (= qkview",
        "   capture time), NOT the script run time. 3h/1d/7d/30d count",
        "   backwards from capture.",
        " - 'Highest' = VDEF MAXIMUM over each window from the MAX RRA;",
        "   if the RRD has no MAX RRA, falls back to MAXIMUM over the",
        "   AVERAGE RRA (peak of averages — matches iHealth GUI plots).",
        " - 'Average' = VDEF AVERAGE over each window from the AVERAGE RRA.",
        " - Windows: 3h / 1d / 7d / 30d.",
        " - CPU reported per-blade/per-core; other metrics aggregate matched DS.",
        " - Throughput (Bits) = the GUI 'Service' throughput series (service",
        "   DS) x8 to bits/s; falls back to bytes_in+bytes_out only if no",
        "   service DS exists in the qkview RRD schema.",
        "",
        "If a metric shows N/A:",
        " - 'no qkview-rrd source found' = the expected RRD xml.gz is not in",
        "   this qkview (e.g., no DNS module on this device).",
        " - 'no DS matched' = adjust ds_match in PERF_METRICS at top of script.",
        " - Run with --discover to dump qkview RRD files and DS names.",
        " - No performance values are ever fabricated.",
    ]
    for i, line in enumerate(notes, 1):
        wsn.cell(row=i, column=1, value=line).font = (
            B if line.endswith(":") else R)
    wsn.column_dimensions["A"].width = 100

    wb.save(out_path)


# =============================================================================
# Modes
# =============================================================================

def run_discover(client: IHealthClient) -> None:
    for qid, label in QKVIEWS.items():
        print(f"\n=== QKView {qid} ({label}) — /var/tmp/qkview-rrd ===")
        files = client.list_files(qid)
        rrds = [(fid, p) for fid, p in files if "/var/tmp/qkview-rrd/" in p]
        if not rrds:
            print("  (none found)"); continue
        with tempfile.TemporaryDirectory() as td:
            for fid, p in sorted(rrds, key=lambda x: x[1]):
                ds: List[str] = []
                xml_gz = os.path.join(td, os.path.basename(p))
                rrd_path = xml_gz + ".rrd"
                if rrd_available() and client.download_file(qid, fid, xml_gz):
                    if rrd_restore_xml_gz(xml_gz, rrd_path):
                        ds = rrd_ds_names(rrd_path)
                print(f"  {os.path.basename(p)}")
                if ds:
                    print(f"      DS: {', '.join(ds)}")


def run_analyze(client: IHealthClient) -> None:
    rows: List[Dict[str, Any]] = []
    for qid, label in QKVIEWS.items():
        print(f"\n=== QKView {qid} ({label}) ===", file=sys.stderr)
        files = client.list_files(qid)
        print(f"  files in qkview: {len(files)}", file=sys.stderr)
        rows.append({
            "label": label,
            "identity":     extract_identity(client, qid),
            "license_prov": extract_license_provisioning(client, qid, files),
            "config":       extract_config_totals(client, qid, files),
            "performance":  extract_performance(client, qid, files),
        })
    build_workbook(rows, OUTPUT_XLSX)
    print(f"\nWrote {OUTPUT_XLSX}")


def main() -> int:
    ap = argparse.ArgumentParser(description="iHealth QKView sizing extractor")
    ap.add_argument("--discover", action="store_true",
                    help="list /var/tmp/qkview-rrd files + DS names per qkview")
    ap.add_argument("--upload", metavar="FILE",
                    help="upload a .qkview, print the new qkview id, exit")
    args = ap.parse_args()

    client = IHealthClient(); client.authenticate()
    if args.upload:
        new_id = client.upload(args.upload)
        print(f"new qkview id: {new_id}")
        return 0
    if not QKVIEWS:
        print("Edit QKVIEWS at top of script.", file=sys.stderr); return 2
    if args.discover:
        run_discover(client)
    else:
        run_analyze(client)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
